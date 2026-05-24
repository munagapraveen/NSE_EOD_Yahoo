# -*- coding: utf-8 -*-
"""
sharpe_screener.py -- NSE Sharpe Ratio Screener
=================================================
Screens NSE EQ + BE stocks using Sharpe Ratio.

PIPELINE:
  1. MCAP snapshot   -- latest close + fundamentals (one row per symbol)
  2. MCAP filter     -- Market Cap > 1000 Cr
  3. Price history   -- load last 253 days (close + volume) for filtered symbols
  4. ROC filter      -- Annual ROC >= 6.5%  (close_today vs close_252d_ago)
  5. Turnover filter -- Median Daily Turnover >= 1 Cr  (median of volume x close
                        over 252 days). Median used instead of Mean so that a few
                        high-volume block deal days do not inflate an illiquid stock.
  6. Sharpe          -- mean(daily_returns) / std(daily_returns, ddof=1)
                        sqrt(252) omitted -- does not affect ranking, only scale
  7. Rank            -- sharpe_6_rank, sharpe_3_rank, Avg_sharpe_6_3_Rank

Usage:
    python sharpe_screener.py
    python sharpe_screener.py --top 100
    python sharpe_screener.py --mcap 500
    python sharpe_screener.py --rf 7.0          # change ROC hurdle %
    python sharpe_screener.py --turnover 5.0    # change turnover floor to 5 Cr
    python sharpe_screener.py --mcap 2000 --top 20

Requirements:
    pip install pandas numpy
"""

import sys
import time
import numpy as np
import pandas as pd
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import (
    get_connection,
    load_indicator_snapshot,
    setup_schema,
    upsert_indicator_rows,
)
from logger import get_logger

log = get_logger(__name__)

# ===========================================================================
# CONFIGURATION
# ===========================================================================

MCAP_FILTER_CR       = 1000   # minimum market cap in Rs. Crore
ROC_ANNUAL_FILTER    = 6.5    # minimum annual ROC % (risk-free rate hurdle)
TURNOVER_FILTER_CR   = 1.0    # minimum MEDIAN daily turnover in Rs. Crore
TOP_N                = 50     # how many top stocks to display
# Output filename is set dynamically as <latest_date>.xlsx
ROC_3M_FILTER        = 20.0   # filter: 3M ROC > 20%
CIRCUIT_MAX_HITS     = 10     # filter: circuit hits in last 3 months <= 10
CIRCUIT_LOOKBACK     = 63     # 3 months of trading days for circuit detection
CIRCUIT_BANDS        = [5.0, 10.0, 20.0]  # NSE circuit band %
CIRCUIT_TOLERANCE    = 0.025  # tolerance around each band
TRADING_DAYS_52W     = 252    # 52-week high lookback
TRADING_DAYS_6M_ROC  = 126    # 6-month ROC window
DMA_20               = 20     # 20-day moving average
DMA_50               = 50     # 50-day moving average
DMA_100              = 100    # 100-day moving average
DMA_200              = 200    # 200-day moving average

DAYS_TO_LOAD         = 253    # 252 returns = 253 prices (covers full 1-year window)
TRADING_DAYS_6M      = 126    # approx 6 months
TRADING_DAYS_3M      = 63     # approx 3 months
TRADING_DAYS_PER_MONTH = 21   # approximate trading days in a month


def months_to_trading_days(months):
    """Convert a month count to approximate trading days."""
    return max(1, int(months) * TRADING_DAYS_PER_MONTH)


# ===========================================================================
# STEP 1 -- MCAP SNAPSHOT
# ===========================================================================

def load_mcap_snapshot(conn, as_of_date=None):
    """
    Loads close + fundamentals for each symbol as of a specific date.
    If as_of_date is None, uses the latest available date in the DB.
    One row per symbol -- very fast query.
    Returns DataFrame and the resolved snapshot_date.
    """
    t0 = time.time()

    if as_of_date:
        # Find the closest available trading date <= as_of_date
        snapshot_date = conn.execute(
            "SELECT MAX(date) FROM eod_data WHERE date <= ?",
            (as_of_date,)
        ).fetchone()[0]
        if not snapshot_date:
            raise ValueError(
                f"No data found on or before {as_of_date}. "
                f"Check the date or run downloader.py first."
            )
        log.info(f"Step 1 -- Loading MCAP snapshot as of {snapshot_date} ...")
    else:
        snapshot_date = conn.execute(
            "SELECT MAX(date) FROM eod_data"
        ).fetchone()[0]
        if not snapshot_date:
            raise ValueError("No data in eod_data. Run downloader.py first.")
        log.info("Step 1 -- Loading MCAP snapshot (latest date) ...")

    df = pd.read_sql("""
        SELECT
            e.symbol,
            e.company_name,
            e.isin,
            f.sector,
            f.industry,
            m.shares_outstanding,
            m.market_cap_cr,
            e.close,
            i.ma_20 AS dma_20,
            i.ma_50 AS dma_50,
            i.ma_100 AS dma_100,
            i.ma_200 AS dma_200
        FROM eod_data e
        LEFT JOIN marketcap m
            ON e.symbol = m.symbol AND e.date = m.date
        LEFT JOIN fundamentals f ON e.symbol = f.symbol
        LEFT JOIN indicators i
            ON e.symbol = i.symbol AND e.date = i.date
        WHERE e.date = ?
          AND e.close > 0
    """, conn, params=[snapshot_date])

    log.info(
        f"  Snapshot loaded : {len(df):,} symbols  "
        f"({time.time() - t0:.1f}s)"
    )
    return df, snapshot_date


# ===========================================================================
# STEP 2 -- MCAP FILTER
# ===========================================================================

def apply_mcap_filter(snapshot_df, mcap_filter):
    """Filters to symbols with MCAP >= mcap_filter Cr."""
    filtered = snapshot_df[
        snapshot_df["market_cap_cr"].notna() &
        (snapshot_df["market_cap_cr"] >= mcap_filter)
    ].copy()

    log.info(
        f"Step 2 -- MCAP filter > Rs.{mcap_filter:,} Cr : "
        f"{len(filtered):,} symbols  "
        f"(from {len(snapshot_df):,})"
    )

    if filtered.empty:
        log.warning("No stocks passed MCAP filter.")
        log.warning("Run marketcap.py --fetch first to populate fundamentals.")

    return filtered


# ===========================================================================
# STEP 3 -- LOAD PRICE HISTORY (filtered symbols, limited rows)
# ===========================================================================

def load_price_history(conn, symbols, days=DAYS_TO_LOAD, as_of_date=None):
    """
    Loads last `days` rows of close + volume for filtered symbols only.
    If as_of_date is provided, only rows on or before that date are used.
    Uses ROW_NUMBER window function to efficiently slice last N rows per symbol.
    Batches symbols to stay within SQLite's IN clause parameter limit.
    """
    t0 = time.time()
    n  = len(symbols)
    date_label = f"up to {as_of_date}" if as_of_date else "latest"
    log.info(
        f"Step 3 -- Loading last {days} days price history "
        f"for {n:,} symbols ({date_label}) ..."
    )

    BATCH      = 900
    all_frames = []

    for i in range(0, n, BATCH):
        batch   = symbols[i: i + BATCH]
        holders = ",".join(["?"] * len(batch))

        # Add date filter when running on historical date
        date_filter = f"AND date <= '{as_of_date}'" if as_of_date else ""

        df_batch = pd.read_sql(f"""
            SELECT symbol, date, close, volume
            FROM (
                SELECT
                    symbol, date, close, volume,
                    ROW_NUMBER() OVER (
                        PARTITION BY symbol
                        ORDER BY date DESC
                    ) AS rn
                FROM eod_data
                WHERE symbol IN ({holders})
                  AND close > 0
                  {date_filter}
            )
            WHERE rn <= {days}
        """, conn, params=batch)

        all_frames.append(df_batch)

    prices_df = (
        pd.concat(all_frames, ignore_index=True)
        if all_frames else pd.DataFrame()
    )

    # Sort in pandas -- faster than ORDER BY in SQL on large result sets
    prices_df.sort_values(["symbol", "date"], inplace=True)
    prices_df.reset_index(drop=True, inplace=True)

    log.info(
        f"  Price history loaded : {len(prices_df):,} rows  "
        f"({time.time() - t0:.1f}s)"
    )
    return prices_df


# ===========================================================================
# STEP 4 -- ANNUAL ROC FILTER
# ===========================================================================

def compute_and_filter_roc(prices_df, roc_filter):
    """
    Annual ROC = (close_today - close_252d_ago) / close_252d_ago * 100

    Requires exactly DAYS_TO_LOAD (253) rows -- stocks with fewer rows
    are newly listed (< 1 year history) and are excluded completely.
    This prevents a 3-month-old IPO from using its IPO price as the
    252-day-ago reference and inflating its ROC figure.

        close_today    = prices[-1]  (most recent)
        close_252d_ago = prices[0]   (oldest -- exactly 252 trading days ago)

    Stocks below roc_filter % are excluded.
    Returns DataFrame: symbol, ROC_annual
    """
    log.info(f"Step 4 -- Annual ROC filter >= {roc_filter}% ...")

    grouped  = prices_df.groupby("symbol")["close"]
    roc_list = []

    for symbol, closes in grouped:
        arr = closes.values.astype(float)

        # Must have a full 1-year window (253 prices = 252 returns)
        # Stocks with fewer rows are newly listed -- skip them entirely
        if len(arr) < DAYS_TO_LOAD:
            continue

        close_today  = arr[-1]
        close_yr_ago = arr[0]   # exactly 252 trading days ago

        if close_yr_ago <= 0:
            continue

        roc = (close_today - close_yr_ago) / close_yr_ago * 100
        roc_list.append({"symbol": symbol, "ROC_annual": round(roc, 2)})

    roc_df   = pd.DataFrame(roc_list)
    filtered = roc_df[roc_df["ROC_annual"] >= roc_filter].copy()

    total_mcap_symbols = prices_df["symbol"].nunique()
    no_history = total_mcap_symbols - len(roc_df)
    below_roc  = len(roc_df) - len(filtered)

    log.info(f"  Skipped (< 1yr history) : {no_history:,} newly listed stocks")
    log.info(f"  Skipped (ROC < {roc_filter}%)  : {below_roc:,} stocks")
    log.info(f"  Passed ROC filter       : {len(filtered):,} symbols")
    return filtered


# ===========================================================================
# STEP 5 -- MEDIAN DAILY TURNOVER FILTER
# ===========================================================================

def compute_and_filter_turnover(prices_df, symbols, turnover_filter_cr=TURNOVER_FILTER_CR):
    """
    Computes Median Daily Turnover for each symbol over last 252 days.

        Daily Turnover (Rs.) = volume x close
        Median Daily Turnover = median(daily_turnover over 252 days)

    WHY MEDIAN not Mean:
        A few block deal days or index rebalancing events can spike volume,
        inflating the Mean and making an illiquid stock appear liquid.
        Median reflects the TYPICAL daily liquidity, ignoring outlier days.

    Filter: Median Daily Turnover >= turnover_filter_cr Cr
    Removes illiquid stocks that are difficult to enter/exit in real size.

    Returns DataFrame: symbol, median_turnover_cr
    """
    log.info(
        f"Step 5 -- Median Daily Turnover filter "
        f">= Rs.{turnover_filter_cr} Cr ..."
    )

    subset  = prices_df[prices_df["symbol"].isin(symbols)].copy()
    grouped = subset.groupby("symbol")

    results = []

    for symbol, grp in grouped:
        grp = grp.sort_values("date")

        # Use last 252 rows for the turnover window
        window = grp.tail(252)

        # Require full 252-day history -- newly listed stocks with
        # fewer rows are excluded to avoid misleading turnover readings
        if len(window) < 252:
            continue

        closes  = window["close"].values.astype(float)
        volumes = window["volume"].values.astype(float)

        # Daily turnover in Rs.
        daily_turnover_rs = closes * volumes

        # MEDIAN -- not mean
        median_turnover_cr = round(np.median(daily_turnover_rs) / 1e7, 4)

        results.append({
            "symbol":             symbol,
            "median_turnover_cr": median_turnover_cr,
        })

    turnover_df = pd.DataFrame(results)

    if turnover_df.empty:
        return pd.DataFrame()

    filtered = turnover_df[
        turnover_df["median_turnover_cr"] >= turnover_filter_cr
    ].copy()

    excluded = len(turnover_df) - len(filtered)
    log.info(
        f"  Median Turnover >= Rs.{turnover_filter_cr} Cr : "
        f"{len(filtered):,} symbols  "
        f"(excluded {excluded:,} illiquid stocks)"
    )
    return filtered


# ===========================================================================
# STEP 6 -- SHARPE COMPUTATION (vectorised numpy)
# ===========================================================================

def compute_sharpe_vectorised(prices_df, symbols, long_months=6, short_months=3):
    """
    Computes long-window and short-window Sharpe ratios for all symbols.

    Formula (sqrt(252) omitted -- ranking is identical without it):
        daily_returns = diff(close) / close[:-1]  (simple returns)
        sharpe        = mean(daily_returns) / std(daily_returns, ddof=1)

    All computation done in numpy arrays.
    """
    long_days = months_to_trading_days(long_months)
    short_days = months_to_trading_days(short_months)
    log.info(
        f"Step 6 -- Computing Sharpe ratios for {len(symbols):,} symbols "
        f"({long_months}M / {short_months}M) ..."
    )

    t0      = time.time()
    results = []
    skipped = 0

    subset  = prices_df[prices_df["symbol"].isin(symbols)]
    grouped = subset.groupby("symbol")["close"]

    for symbol, closes in grouped:
        prices = closes.values.astype(float)

        # -- Long-window Sharpe --
        if len(prices) >= DAYS_TO_LOAD and len(prices) >= long_days + 1:
            w6 = prices[-(long_days + 1):]
            r6 = np.diff(w6) / w6[:-1]
            s6 = r6.std(ddof=1)
            sharpe_6 = round(r6.mean() / s6, 4) if s6 > 0 else None
        else:
            sharpe_6 = None

        # -- Short-window Sharpe --
        if len(prices) >= DAYS_TO_LOAD and len(prices) >= short_days + 1:
            w3 = prices[-(short_days + 1):]
            r3 = np.diff(w3) / w3[:-1]
            s3 = r3.std(ddof=1)
            sharpe_3 = round(r3.mean() / s3, 4) if s3 > 0 else None
        else:
            sharpe_3 = None

        if sharpe_6 is None and sharpe_3 is None:
            skipped += 1
            continue

        # -- Short-window ROC --
        roc_3 = None
        if len(prices) >= short_days + 1:
            p_now   = prices[-1]
            p_3m    = prices[-(short_days + 1)]
            roc_3   = round((p_now - p_3m) / p_3m * 100, 2) if p_3m > 0 else None

        # -- 52-Week High --
        week_52_high = None
        if len(prices) >= TRADING_DAYS_52W:
            week_52_high = round(float(np.max(prices[-TRADING_DAYS_52W:])), 2)

        # -- Long-window ROC --
        roc_6 = None
        if len(prices) >= long_days + 1:
            p_6m  = prices[-(long_days + 1)]
            roc_6 = round((prices[-1] - p_6m) / p_6m * 100, 2) if p_6m > 0 else None

        # -- Away from 52W High (negative value = below peak) --
        close_today = prices[-1]
        away_52wh = None
        if week_52_high and week_52_high > 0:
            away_52wh = round((close_today - week_52_high) / week_52_high * 100, 2)

        results.append({
            "symbol":       symbol,
            "sharpe_6":     sharpe_6,
            "sharpe_3":     sharpe_3,
            "ROC_6":        roc_6,
            "ROC_3":        roc_3,
            "week_52_high": week_52_high,
            "away_52wh":    away_52wh,
        })

    log.info(
        f"  Sharpe computed : {len(results):,}  "
        f"skipped : {skipped:,}  "
        f"({time.time() - t0:.1f}s)"
    )
    return pd.DataFrame(results)


def compute_snapshot_indicators(prices_df, snapshot_date):
    """Compute and persist MA rows for one snapshot date."""
    records = []
    subset = prices_df[prices_df["date"] <= snapshot_date].copy()
    grouped = subset.groupby("symbol")["close"]

    for symbol, closes in grouped:
        prices = closes.values.astype(float)
        records.append({
            "symbol": symbol,
            "date": snapshot_date,
            "ma_5": round(float(np.mean(prices[-5:])), 2) if len(prices) >= 5 else None,
            "ma_10": round(float(np.mean(prices[-10:])), 2) if len(prices) >= 10 else None,
            "ma_20": round(float(np.mean(prices[-DMA_20:])), 2) if len(prices) >= DMA_20 else None,
            "ma_50": round(float(np.mean(prices[-DMA_50:])), 2) if len(prices) >= DMA_50 else None,
            "ma_100": round(float(np.mean(prices[-DMA_100:])), 2) if len(prices) >= DMA_100 else None,
            "ma_200": round(float(np.mean(prices[-DMA_200:])), 2) if len(prices) >= DMA_200 else None,
        })

    return pd.DataFrame(records)


# ===========================================================================
# CIRCUIT HIT COMPUTATION (reuses circuit_analyser logic inline)
# ===========================================================================

def compute_circuit_hits(prices_df, symbols,
                         lookback_days=CIRCUIT_LOOKBACK,
                         bands=CIRCUIT_BANDS,
                         tol=CIRCUIT_TOLERANCE):
    """
    Counts circuit hits per symbol over last lookback_days trading days.
    Uses same detection logic as circuit_analyser.py:
        daily_return = (close - prev_close) / prev_close * 100
        Circuit hit  = return within +-tol of 5%, 10%, or 20%

    Returns DataFrame: symbol, total_circuit_hits_3m
    """
    log.info(
        f"Computing circuit hits (last {lookback_days} days) "
        f"for {len(symbols):,} symbols ..."
    )

    subset  = prices_df[prices_df["symbol"].isin(symbols)]
    grouped = subset.groupby("symbol")["close"]
    results = []

    for symbol, closes in grouped:
        arr = closes.values.astype(float)

        # Take last lookback_days + 1 prices to get lookback_days returns
        window = arr[-(lookback_days + 1):]
        if len(window) < 2:
            continue

        prev_c = window[:-1]
        curr_c = window[1:]

        valid   = prev_c > 0
        returns = np.where(
            valid,
            (curr_c - prev_c) / prev_c * 100,
            np.nan
        )

        hits = 0
        for ret in returns:
            if np.isnan(ret):
                continue
            for band in bands:
                if (band - tol) <= ret <= (band + tol):
                    hits += 1
                    break
                if -(band + tol) <= ret <= -(band - tol):
                    hits += 1
                    break

        results.append({"symbol": symbol, "total_circuit_hits_3m": hits})

    df = pd.DataFrame(results)
    log.info(
        f"  Circuit hits computed : {len(df):,} symbols  |  "
        f"{(df['total_circuit_hits_3m'] > 0).sum():,} had at least one hit"
    )
    return df


# ===========================================================================
# STEP 7 -- RANKING
# ===========================================================================

def rank_stocks(df):
    """
    sharpe_6_rank       : rank by sharpe_6 descending (1 = best)
    sharpe_3_rank       : rank by sharpe_3 descending
    Avg_sharpe_6_3_Rank : sum of both ranks, sorted ascending
    """
    df = df.dropna(subset=["sharpe_6", "sharpe_3"]).copy()

    df["sharpe_6_rank"] = (
        df["sharpe_6"]
        .rank(ascending=False, method="first")
        .astype(int)
    )
    df["sharpe_3_rank"] = (
        df["sharpe_3"]
        .rank(ascending=False, method="first")
        .astype(int)
    )
    df["Avg_sharpe_6_3_Rank"] = (
        df["sharpe_6_rank"] + df["sharpe_3_rank"]
    )

    df.sort_values("Avg_sharpe_6_3_Rank", ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


# ===========================================================================
# SCREENER PIPELINE
# ===========================================================================

def run_screener(
    mcap_filter=MCAP_FILTER_CR,
    roc_filter=ROC_ANNUAL_FILTER,
    turnover_filter=TURNOVER_FILTER_CR,
    as_of_date=None,
    long_months=6,
    short_months=3,
):
    """
    Full Sharpe screener pipeline. Returns ranked DataFrame.

    Args:
        as_of_date : str in YYYY-MM-DD format, or None for today.
                     When provided, the screener runs as if that date
                     is today -- using only data available up to that date.
                     Useful for historical analysis and validation.
    """

    if long_months <= 0 or short_months <= 0:
        raise ValueError("Sharpe month windows must be positive integers.")
    if long_months > 12 or short_months > 12:
        raise ValueError("Sharpe month windows cannot exceed 12.")

    t_start = time.time()
    long_days = months_to_trading_days(long_months)
    short_days = months_to_trading_days(short_months)
    days_to_load = max(DAYS_TO_LOAD, long_days + 1, short_days + 1)

    log.info("")
    log.info("=" * 60)
    log.info("NSE SHARPE RATIO SCREENER")
    log.info("=" * 60)
    if as_of_date:
        log.info(f"  Mode            : HISTORICAL  (as of {as_of_date})")
    else:
        log.info(f"  Mode            : LIVE  (latest data)")
    log.info(f"  MCAP filter     : > Rs.{mcap_filter:,} Cr")
    log.info(f"  Annual ROC      : >= {roc_filter}%")
    log.info(f"  Median Turnover : >= Rs.{turnover_filter} Cr / day")
    log.info(f"  Sharpe windows  : {long_months}M and {short_months}M")
    log.info(f"  Sharpe formula  : mean(r) / std(r)  [no sqrt(252)]")
    log.info(f"  Price window    : last {days_to_load} trading days")
    log.info("")

    with get_connection() as conn:
        setup_schema(conn)

        # Step 1 -- MCAP snapshot
        snapshot_df, latest_date = load_mcap_snapshot(conn, as_of_date=as_of_date)

        # Step 2 -- MCAP filter
        mcap_df = apply_mcap_filter(snapshot_df, mcap_filter)
        if mcap_df.empty:
            return pd.DataFrame(), None

        # Step 3 -- Price history for filtered symbols only
        prices_df = load_price_history(
            conn, mcap_df["symbol"].tolist(),
            days=days_to_load, as_of_date=as_of_date
        )

    # All remaining steps run in memory -- DB connection closed
    if prices_df.empty:
        log.warning("No price history found.")
        return pd.DataFrame(), None

    indicator_calc_df = compute_snapshot_indicators(prices_df, latest_date)
    indicator_df = pd.DataFrame(columns=["symbol", "dma_20", "dma_50", "dma_100", "dma_200"])
    if not indicator_calc_df.empty:
        with get_connection() as conn:
            upsert_indicator_rows(conn, indicator_calc_df.to_dict("records"))
            indicator_snapshot = load_indicator_snapshot(conn, latest_date)
        if not indicator_snapshot.empty:
            indicator_df = indicator_snapshot.rename(columns={
                "ma_20": "dma_20",
                "ma_50": "dma_50",
                "ma_100": "dma_100",
                "ma_200": "dma_200",
            })[["symbol", "dma_20", "dma_50", "dma_100", "dma_200"]]

    # Step 4 -- Annual ROC filter
    roc_df = compute_and_filter_roc(prices_df, roc_filter)
    if roc_df.empty:
        log.warning(f"No stocks passed Annual ROC >= {roc_filter}% filter.")
        return pd.DataFrame(), None

    # Step 5 -- Median Daily Turnover filter
    turnover_df = compute_and_filter_turnover(
        prices_df, roc_df["symbol"].tolist(), turnover_filter
    )
    if turnover_df.empty:
        log.warning(
            f"No stocks passed Median Turnover >= Rs.{turnover_filter} Cr filter."
        )
        return pd.DataFrame(), None

    # Step 6 -- Sharpe computation on fully filtered set
    sharpe_df = compute_sharpe_vectorised(
        prices_df,
        turnover_df["symbol"].tolist(),
        long_months=long_months,
        short_months=short_months,
    )
    if sharpe_df.empty:
        log.warning("No stocks had sufficient history for Sharpe.")
        return pd.DataFrame(), None

    # Step 7 -- Circuit hits computation
    log.info("Step 7 -- Computing circuit hits ...")
    circuit_df = compute_circuit_hits(
        prices_df, sharpe_df["symbol"].tolist()
    )

    # Step 8 -- Merge all results
    log.info("Step 8 -- Merging and ranking ...")
    snapshot_base = mcap_df.drop(
        columns=[col for col in ["dma_20", "dma_50", "dma_100", "dma_200"] if col in mcap_df.columns],
        errors="ignore",
    )
    result = (
        snapshot_base
        .merge(indicator_df, on="symbol", how="left")
        .merge(roc_df,      on="symbol", how="inner")
        .merge(turnover_df, on="symbol", how="inner")
        .merge(sharpe_df,   on="symbol", how="inner")
        .merge(circuit_df,  on="symbol", how="left")
    )
    # Fill NaN circuit hits with 0 (symbol had no hits)
    result["total_circuit_hits_3m"] = (
        result["total_circuit_hits_3m"].fillna(0).astype(int)
    )
    log.info(f"  Pool before ranking : {len(result):,} stocks")

    result = rank_stocks(result)
    result.attrs["long_months"] = int(long_months)
    result.attrs["short_months"] = int(short_months)

    output_cols = [
        # Requested column order
        "symbol", "company_name",
        "close", "dma_20", "dma_50", "dma_100", "dma_200",
        "away_52wh",
        "Avg_sharpe_6_3_Rank",
        "sharpe_6", "sharpe_3",
        "ROC_6", "ROC_3",
        # Additional context columns
        "week_52_high", "market_cap_cr", "ROC_annual",
        "median_turnover_cr",
        "sharpe_6_rank", "sharpe_3_rank",
        "isin", "sector", "industry",
    ]
    result = result[[c for c in output_cols if c in result.columns]]

    elapsed = time.time() - t_start
    log.info(f"  Final ranked list : {len(result):,} stocks")
    log.info(f"  Total time        : {elapsed:.1f}s")
    log.info("")
    return result, latest_date


# ===========================================================================
# DISPLAY
# ===========================================================================


# ===========================================================================
# EXCEL EXPORT
# ===========================================================================

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")  # dark blue
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")  # light blue
FILTER_FILL   = PatternFill("solid", start_color="1A5C38")  # dark green
ALT_FILL2     = PatternFill("solid", start_color="D4EDDA")  # light green
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Arial", size=9)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
THIN          = Side(style="thin", color="B0B0B0")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    "symbol":             12,
    "company_name":       28,
    "close":              10,
    "dma_20":             10,
    "dma_50":             10,
    "dma_100":            10,
    "dma_200":            10,
    "away_52wh":          15,
    "Avg_sharpe_6_3_Rank":14,
    "sharpe_6":           11,
    "sharpe_3":           11,
    "ROC_6":              10,
    "ROC_3":              10,
    "week_52_high":       13,
    "market_cap_cr":      14,
    "ROC_annual":         12,
    "median_turnover_cr": 16,
    "total_circuit_hits_3m": 18,
    "sharpe_6_rank":      13,
    "sharpe_3_rank":      13,
    "isin":               14,
    "sector":             18,
    "industry":           22,
}

FRIENDLY_HEADERS = {
    "symbol":             "Symbol",
    "company_name":       "Company Name",
    "close":              "Close (Rs.)",
    "dma_20":             "20 DMA",
    "dma_50":             "50 DMA",
    "dma_100":            "100 DMA",
    "dma_200":            "200 DMA",
    "away_52wh":          "Away from 52WH %",
    "total_circuit_hits_3m":"Circuit Hits (3M)",
    "Avg_sharpe_6_3_Rank":"Avg Sharpe Rank",
    "sharpe_6":           "Sharpe 6M",
    "sharpe_3":           "Sharpe 3M",
    "ROC_6":              "6M ROC %",
    "ROC_3":              "3M ROC %",
    "week_52_high":       "52W High (Rs.)",
    "market_cap_cr":      "MCAP (Cr)",
    "ROC_annual":         "Annual ROC %",
    "median_turnover_cr": "Med. Turnover (Cr)",
    "sharpe_6_rank":      "Sharpe 6M Rank",
    "sharpe_3_rank":      "Sharpe 3M Rank",
    "isin":               "ISIN",
    "sector":             "Sector",
    "industry":           "Industry",
}

NUM_COLS   = {"close", "dma_20", "dma_50", "dma_100", "dma_200",
              "week_52_high", "market_cap_cr", "median_turnover_cr",
              "sharpe_6", "sharpe_3"}
PCT_COLS   = {"ROC_annual", "ROC_3", "ROC_6", "away_52wh"}
INT_COLS   = {"sharpe_6_rank", "sharpe_3_rank", "Avg_sharpe_6_3_Rank", "total_circuit_hits_3m"}


def _write_sheet(ws, df, title, header_fill, alt_fill, friendly_headers=None):
    """Writes a DataFrame to an openpyxl worksheet with formatting."""
    cols = list(df.columns)
    today_str = datetime.today().strftime("%d %b %Y")
    headers = friendly_headers or FRIENDLY_HEADERS

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1,
                         value=f"{title}  |  Screened: {today_str}")
    title_cell.font      = Font(name="Arial", bold=True,
                                color="FFFFFF", size=11)
    title_cell.fill      = header_fill
    title_cell.alignment = CENTER

    # Header row
    for ci, col in enumerate(cols, start=1):
        cell           = ws.cell(row=2, column=ci,
                                 value=headers.get(col, col))
        cell.font      = HEADER_FONT
        cell.fill      = header_fill
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(cols, start=1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            cell.fill   = fill

            if col in PCT_COLS:
                cell.value        = val / 100 if pd.notna(val) else None
                cell.number_format = "0.00%"
                cell.alignment    = CENTER
                cell.font         = DATA_FONT
            elif col in NUM_COLS:
                cell.value        = round(float(val), 2) if pd.notna(val) else None
                cell.number_format = "#,##0.00"
                cell.alignment    = CENTER
                cell.font         = DATA_FONT
            elif col in INT_COLS:
                cell.value        = int(val) if pd.notna(val) else None
                cell.number_format = "0"
                cell.alignment    = CENTER
                cell.font         = DATA_FONT
            else:
                cell.value     = str(val) if pd.notna(val) else ""
                cell.alignment = LEFT
                cell.font      = DATA_FONT

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Auto-filter on header row
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(len(cols))}2"
    )

    # Row height
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, 3 + len(df)):
        ws.row_dimensions[r].height = 15


def apply_filtered_workbook(df):
    """
    Applies three filters to produce the filtered dataset:
        a) ROC_3 > 20%                    (strong 3-month momentum)
        b) close >= 52W_High * 0.75       (within 25% of 52-week high)
        c) total_circuit_hits_3m <= 10    (not excessively circuit-prone)
    All three conditions must pass.
    Returns filtered DataFrame.
    """
    filtered = df.copy()

    # Filter a: 3M ROC > 20%
    if "ROC_3" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["ROC_3"].notna() &
            (filtered["ROC_3"] > ROC_3M_FILTER)
        ]
        log.info(
            f"  Filter a (ROC_3 > {ROC_3M_FILTER}%) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    # Filter b: close >= 52W_High * 0.75 (within 25% of 52-week high)
    if "week_52_high" in filtered.columns and "close" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["week_52_high"].notna() &
            (filtered["close"] >= filtered["week_52_high"] * 0.75)
        ]
        log.info(
            f"  Filter b (Close >= 52WH x 75%) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    # Filter c: circuit hits in last 3 months <= CIRCUIT_MAX_HITS
    if "total_circuit_hits_3m" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["total_circuit_hits_3m"] <= CIRCUIT_MAX_HITS
        ]
        log.info(
            f"  Filter c (Circuit hits <= {CIRCUIT_MAX_HITS}) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    return filtered.reset_index(drop=True)


def export_to_excel(result, latest_date):
    """
    Exports both workbooks as two sheets inside a single Excel file.
    File is named <latest_date>.xlsx  e.g. 2024-04-17.xlsx

    Sheet 1 -- All Stocks  : full ranked output
    Sheet 2 -- Filtered    : ROC_3 > 20% AND close >= 52W_High*0.75 AND circuit_hits <= 10
    """
    base      = Path(__file__).parent
    filename  = f"{latest_date}.xlsx"
    out_path  = base / filename
    long_months = int(result.attrs.get("long_months", 6))
    short_months = int(result.attrs.get("short_months", 3))

    friendly_headers = dict(FRIENDLY_HEADERS)
    friendly_headers["sharpe_6"] = f"Sharpe {long_months}M"
    friendly_headers["sharpe_3"] = f"Sharpe {short_months}M"
    friendly_headers["ROC_6"] = f"{long_months}M ROC %"
    friendly_headers["ROC_3"] = f"{short_months}M ROC %"
    friendly_headers["sharpe_6_rank"] = f"Sharpe {long_months}M Rank"
    friendly_headers["sharpe_3_rank"] = f"Sharpe {short_months}M Rank"
    friendly_headers["Avg_sharpe_6_3_Rank"] = (
        f"Avg Sharpe {long_months}M/{short_months}M Rank"
    )

    wb = Workbook()

    # ---- Sheet 1: All ranked stocks ----
    ws1       = wb.active
    ws1.title = "All Stocks"
    _write_sheet(
        ws1, result,
        title="NSE Sharpe Screener -- All Ranked Stocks",
        header_fill=HEADER_FILL,
        alt_fill=ALT_FILL,
        friendly_headers=friendly_headers,
    )
    log.info(f"  Sheet 1 written : All Stocks  ({len(result):,} rows)")

    # ---- Sheet 2: Filtered ----
    filtered  = apply_filtered_workbook(result)
    ws2       = wb.create_sheet(title="Filtered")

    if filtered.empty:
        log.warning(
            "  No stocks passed filtered sheet criteria "
            f"(3M ROC > {ROC_3M_FILTER}% | Close >= 52WH x 75% | Circuit Hits <= {CIRCUIT_MAX_HITS})"
        )
        ws2.cell(row=1, column=1,
                 value="No stocks matched the filter criteria.")
    else:
        _write_sheet(
            ws2, filtered,
            title=(
                f"Filtered  |  3M ROC > {ROC_3M_FILTER}%  |  "
                f"Close >= 52WH x 75%  |  Circuit Hits <= {CIRCUIT_MAX_HITS}"
            ),
            header_fill=FILTER_FILL,
            alt_fill=ALT_FILL2,
            friendly_headers=friendly_headers,
        )
        log.info(
            f"  Sheet 2 written : Filtered  ({len(filtered):,} rows)"
        )

    wb.save(str(out_path))
    log.info(f"Excel file saved : {out_path}")
    log.info(
        f"  Sheet 1 : All Stocks ({len(result):,} stocks)")
    if not filtered.empty:
        log.info(
            f"  Sheet 2 : Filtered  ({len(filtered):,} stocks)  "
            f"[3M ROC > {ROC_3M_FILTER}% | Close >= 52WH x 75%]"
        )


def print_results(df, top_n=TOP_N, roc_filter=ROC_ANNUAL_FILTER,
                  turnover_filter=TURNOVER_FILTER_CR):
    """Prints formatted top N results to terminal."""
    if df.empty:
        print("No results to display.")
        return

    long_months = int(df.attrs.get("long_months", 6))
    short_months = int(df.attrs.get("short_months", 3))
    display = df.head(top_n).copy()
    display.index = range(1, len(display) + 1)

    for col in ["ROC_annual", "ROC_6", "ROC_3", "away_52wh"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:+.2f}%" if x is not None and str(x) != "nan" else ""
            )
    for col in ["sharpe_6", "sharpe_3"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:+.4f}" if x is not None and str(x) != "nan" else ""
            )
    for col in ["close", "dma_20", "dma_50", "dma_100", "dma_200"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:,.2f}" if x is not None and str(x) != "nan" else ""
            )
    if "market_cap_cr" in display.columns:
        display["market_cap_cr"] = display["market_cap_cr"].apply(
            lambda x: f"Rs.{x:,.0f} Cr" if x == x and x is not None else "N/A"
        )
    if "median_turnover_cr" in display.columns:
        display["median_turnover_cr"] = display["median_turnover_cr"].apply(
            lambda x: f"Rs.{x:.2f} Cr" if x == x and x is not None else "N/A"
        )

    cols = [
        "symbol", "company_name", "market_cap_cr",
        "ROC_annual", "median_turnover_cr",
        "sharpe_6", "sharpe_6_rank",
        "sharpe_3", "sharpe_3_rank",
        "Avg_sharpe_6_3_Rank",
    ]
    cols = [c for c in cols if c in display.columns]

    print(f"\n{'='*100}")
    print(
        f"  TOP {top_n} SHARPE RATIO STOCKS  |  "
        f"Screened: {datetime.today().strftime('%d %b %Y')}"
    )
    print(f"{'='*100}")
    print(display[cols].to_string())
    print(f"{'='*100}")
    print(
        f"  Filters : MCAP > Rs.{MCAP_FILTER_CR:,} Cr  |  "
        f"Annual ROC >= {roc_filter}%  |  "
        f"Median Turnover >= Rs.{turnover_filter} Cr"
    )
    print(f"  Windows : Sharpe {long_months}M and {short_months}M")
    print(f"  Formula : Sharpe = mean(daily_returns) / std(daily_returns)")
    print(f"  Note    : Median turnover used (not mean) to ignore block deal spikes")
    print(f"            sqrt(252) omitted -- ranking is identical\n")


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    args = sys.argv[1:]

    mcap_filter     = MCAP_FILTER_CR
    roc_filter      = ROC_ANNUAL_FILTER
    turnover_filter = TURNOVER_FILTER_CR
    top_n           = TOP_N
    as_of_date      = None
    long_months     = 6
    short_months    = 3

    for i, arg in enumerate(args):
        if arg == "--mcap" and i + 1 < len(args):
            try:
                mcap_filter = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--top" and i + 1 < len(args):
            try:
                top_n = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--rf" and i + 1 < len(args):
            try:
                roc_filter = float(args[i + 1])
            except ValueError:
                pass
        if arg == "--turnover" and i + 1 < len(args):
            try:
                turnover_filter = float(args[i + 1])
            except ValueError:
                pass
        if arg == "--date" and i + 1 < len(args):
            raw = args[i + 1].strip()
            try:
                datetime.strptime(raw, "%Y-%m-%d")
                as_of_date = raw
            except ValueError:
                print(f"Invalid --date format: '{raw}'. Use YYYY-MM-DD.")
                return
        if arg == "--long-months" and i + 1 < len(args):
            try:
                long_months = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--short-months" and i + 1 < len(args):
            try:
                short_months = int(args[i + 1])
            except ValueError:
                pass

    if long_months <= 0 or short_months <= 0:
        print("Invalid Sharpe month values. Use positive integers.")
        return
    if long_months > 12 or short_months > 12:
        print("Invalid Sharpe month values. Maximum allowed is 12.")
        return

    result, latest_date = run_screener(
        mcap_filter=mcap_filter,
        roc_filter=roc_filter,
        turnover_filter=turnover_filter,
        as_of_date=as_of_date,
        long_months=long_months,
        short_months=short_months,
    )

    if result.empty or latest_date is None:
        return

    print_results(result, top_n=top_n,
                  roc_filter=roc_filter,
                  turnover_filter=turnover_filter)

    export_to_excel(result, latest_date)


if __name__ == "__main__":
    main()
    #long_months = int(df.attrs.get("long_months", 6))
    #short_months = int(df.attrs.get("short_months", 3))
