# -*- coding: utf-8 -*-
"""
sharpe_screener.py -- NSE Sharpe Ratio Screener (Yahoo/NSE standalone)
======================================================================
Runs the Sharpe screener against the standalone Yahoo/NSE SQLite DB.
No dependency on Zerodha code, DB, or APIs.
"""

import sys
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import get_connection, setup_schema
from logger import get_logger

log = get_logger(__name__)

MCAP_FILTER_CR = 1000
ROC_ANNUAL_FILTER = 6.5
TURNOVER_FILTER_CR = 1.0
TOP_N = 50
ROC_3M_FILTER = 20.0
CIRCUIT_MAX_HITS = 10
CIRCUIT_LOOKBACK = 63
CIRCUIT_BANDS = [5.0, 10.0, 20.0]
CIRCUIT_TOLERANCE = 0.025
TRADING_DAYS_52W = 252
DAYS_TO_LOAD = 253
TRADING_DAYS_PER_MONTH = 21


def months_to_trading_days(months):
    return max(1, int(months) * TRADING_DAYS_PER_MONTH)


def load_mcap_snapshot(conn, as_of_date=None):
    t0 = time.time()

    if as_of_date:
        snapshot_date = conn.execute(
            "SELECT MAX(date) FROM adjusted_eod_prices WHERE date <= ?",
            (as_of_date,),
        ).fetchone()[0]
        if not snapshot_date:
            raise ValueError(f"No adjusted price data found on or before {as_of_date}.")
        log.info(f"Step 1 -- Loading MCAP snapshot as of {snapshot_date} ...")
    else:
        snapshot_date = conn.execute(
            "SELECT MAX(date) FROM adjusted_eod_prices"
        ).fetchone()[0]
        if not snapshot_date:
            raise ValueError("No data in adjusted_eod_prices. Run download + adjust first.")
        log.info("Step 1 -- Loading MCAP snapshot (latest date) ...")

    df = pd.read_sql("""
        SELECT
            a.symbol,
            s.company_name,
            s.isin,
            s.series,
            a.close,
            m.market_cap_cr,
            m.shares_outstanding,
            i.ma_20  AS dma_20,
            i.ma_50  AS dma_50,
            i.ma_100 AS dma_100,
            i.ma_200 AS dma_200
        FROM adjusted_eod_prices a
        LEFT JOIN symbols s
            ON a.symbol = s.symbol
        LEFT JOIN indicators i
            ON a.symbol = i.symbol AND a.date = i.date
        LEFT JOIN marketcap m
            ON a.symbol = m.symbol AND a.date = m.date
        WHERE a.date = ?
          AND a.close > 0
    """, conn, params=[snapshot_date])

    log.info(
        f"  Snapshot loaded : {len(df):,} symbols  "
        f"({time.time() - t0:.1f}s)"
    )
    return df, snapshot_date


def apply_mcap_filter(snapshot_df, mcap_filter):
    filtered = snapshot_df[
        snapshot_df["market_cap_cr"].notna() &
        (snapshot_df["market_cap_cr"] >= mcap_filter)
    ].copy()

    log.info(
        f"Step 2 -- MCAP filter > Rs.{mcap_filter:,} Cr : "
        f"{len(filtered):,} symbols  "
        f"(from {len(snapshot_df):,})"
    )
    return filtered


def load_price_history(conn, symbols, days=DAYS_TO_LOAD, as_of_date=None):
    t0 = time.time()
    n = len(symbols)
    date_label = f"up to {as_of_date}" if as_of_date else "latest"
    log.info(
        f"Step 3 -- Loading last {days} days adjusted history "
        f"for {n:,} symbols ({date_label}) ..."
    )

    batch_size = 900
    all_frames = []

    for i in range(0, n, batch_size):
        batch = symbols[i:i + batch_size]
        placeholders = ",".join(["?"] * len(batch))
        params = list(batch)
        date_clause = ""
        if as_of_date:
            date_clause = "AND date <= ?"
            params.append(as_of_date)

        df_batch = pd.read_sql(f"""
            SELECT symbol, date, close, volume
            FROM (
                SELECT
                    symbol, date, close, volume,
                    ROW_NUMBER() OVER (
                        PARTITION BY symbol
                        ORDER BY date DESC
                    ) AS rn
                FROM adjusted_eod_prices
                WHERE symbol IN ({placeholders})
                  AND close > 0
                  {date_clause}
            )
            WHERE rn <= {days}
        """, conn, params=params)
        all_frames.append(df_batch)

    prices_df = (
        pd.concat(all_frames, ignore_index=True)
        if all_frames else pd.DataFrame()
    )
    prices_df.sort_values(["symbol", "date"], inplace=True)
    prices_df.reset_index(drop=True, inplace=True)

    log.info(
        f"  Price history loaded : {len(prices_df):,} rows  "
        f"({time.time() - t0:.1f}s)"
    )
    return prices_df


def compute_and_filter_roc(prices_df, roc_filter):
    log.info(f"Step 4 -- Annual ROC filter >= {roc_filter}% ...")

    roc_list = []
    grouped = prices_df.groupby("symbol")["close"]

    for symbol, closes in grouped:
        arr = closes.values.astype(float)
        if len(arr) < DAYS_TO_LOAD:
            continue
        if arr[0] <= 0:
            continue
        roc = (arr[-1] - arr[0]) / arr[0] * 100
        roc_list.append({"symbol": symbol, "ROC_annual": round(roc, 2)})

    roc_df = pd.DataFrame(roc_list)
    if roc_df.empty:
        log.warning("No stocks had full 1-year history for ROC.")
        return roc_df

    filtered = roc_df[roc_df["ROC_annual"] >= roc_filter].copy()
    total = prices_df["symbol"].nunique()
    log.info(f"  Skipped (< 1yr history) : {total - len(roc_df):,} symbols")
    log.info(f"  Skipped (ROC < {roc_filter}%)  : {len(roc_df) - len(filtered):,} stocks")
    log.info(f"  Passed ROC filter       : {len(filtered):,} symbols")
    return filtered


def compute_and_filter_turnover(prices_df, symbols, turnover_filter_cr=TURNOVER_FILTER_CR):
    log.info(
        f"Step 5 -- Median Daily Turnover filter "
        f">= Rs.{turnover_filter_cr} Cr ..."
    )

    subset = prices_df[prices_df["symbol"].isin(symbols)].copy()
    results = []

    for symbol, grp in subset.groupby("symbol"):
        grp = grp.sort_values("date").tail(252)
        if len(grp) < 252:
            continue
        closes = grp["close"].values.astype(float)
        volumes = grp["volume"].values.astype(float)
        median_turnover_cr = round(np.median(closes * volumes) / 1e7, 4)
        results.append({
            "symbol": symbol,
            "median_turnover_cr": median_turnover_cr,
        })

    turnover_df = pd.DataFrame(results)
    if turnover_df.empty:
        return turnover_df

    filtered = turnover_df[
        turnover_df["median_turnover_cr"] >= turnover_filter_cr
    ].copy()
    log.info(
        f"  Median Turnover >= Rs.{turnover_filter_cr} Cr : "
        f"{len(filtered):,} symbols  "
        f"(excluded {len(turnover_df) - len(filtered):,} illiquid stocks)"
    )
    return filtered


def compute_sharpe_vectorised(prices_df, symbols, long_months=6, short_months=3):
    long_days = months_to_trading_days(long_months)
    short_days = months_to_trading_days(short_months)
    log.info(
        f"Step 6 -- Computing Sharpe ratios for {len(symbols):,} symbols "
        f"({long_months}M / {short_months}M) ..."
    )

    t0 = time.time()
    results = []
    skipped = 0
    subset = prices_df[prices_df["symbol"].isin(symbols)]

    for symbol, closes in subset.groupby("symbol")["close"]:
        prices = closes.values.astype(float)

        sharpe_6 = None
        if len(prices) >= DAYS_TO_LOAD and len(prices) >= long_days + 1:
            window = prices[-(long_days + 1):]
            returns = np.diff(window) / window[:-1]
            std = returns.std(ddof=1)
            sharpe_6 = round(returns.mean() / std, 4) if std > 0 else None

        sharpe_3 = None
        if len(prices) >= DAYS_TO_LOAD and len(prices) >= short_days + 1:
            window = prices[-(short_days + 1):]
            returns = np.diff(window) / window[:-1]
            std = returns.std(ddof=1)
            sharpe_3 = round(returns.mean() / std, 4) if std > 0 else None

        if sharpe_6 is None and sharpe_3 is None:
            skipped += 1
            continue

        roc_short = None
        if len(prices) >= short_days + 1 and prices[-(short_days + 1)] > 0:
            base = prices[-(short_days + 1)]
            roc_short = round((prices[-1] - base) / base * 100, 2)

        roc_long = None
        if len(prices) >= long_days + 1 and prices[-(long_days + 1)] > 0:
            base = prices[-(long_days + 1)]
            roc_long = round((prices[-1] - base) / base * 100, 2)

        week_52_high = None
        if len(prices) >= TRADING_DAYS_52W:
            week_52_high = round(float(np.max(prices[-TRADING_DAYS_52W:])), 2)

        away_52wh = None
        if week_52_high and week_52_high > 0:
            away_52wh = round((prices[-1] - week_52_high) / week_52_high * 100, 2)

        results.append({
            "symbol": symbol,
            "sharpe_6": sharpe_6,
            "sharpe_3": sharpe_3,
            "ROC_6": roc_long,
            "ROC_3": roc_short,
            "week_52_high": week_52_high,
            "away_52wh": away_52wh,
        })

    log.info(
        f"  Sharpe computed : {len(results):,}  "
        f"skipped : {skipped:,}  "
        f"({time.time() - t0:.1f}s)"
    )
    return pd.DataFrame(results)


def compute_circuit_hits(prices_df, symbols,
                         lookback_days=CIRCUIT_LOOKBACK,
                         bands=CIRCUIT_BANDS,
                         tol=CIRCUIT_TOLERANCE):
    log.info(
        f"Step 7 -- Computing circuit hits (last {lookback_days} days) "
        f"for {len(symbols):,} symbols ..."
    )

    results = []
    subset = prices_df[prices_df["symbol"].isin(symbols)]
    for symbol, closes in subset.groupby("symbol")["close"]:
        arr = closes.values.astype(float)
        window = arr[-(lookback_days + 1):]
        if len(window) < 2:
            continue
        prev_c = window[:-1]
        curr_c = window[1:]
        valid = prev_c > 0
        returns = np.where(valid, (curr_c - prev_c) / prev_c * 100, np.nan)

        hits = 0
        for ret in returns:
            if np.isnan(ret):
                continue
            for band in bands:
                if (band - tol) <= ret <= (band + tol):
                    hits += 1
                    break
                if -(band + tol) <= ret <= -(band - tol):
                    hits += 1
                    break
        results.append({"symbol": symbol, "total_circuit_hits_3m": hits})

    df = pd.DataFrame(results)
    if not df.empty:
        log.info(
            f"  Circuit hits computed : {len(df):,} symbols  |  "
            f"{(df['total_circuit_hits_3m'] > 0).sum():,} had at least one hit"
        )
    return df


def rank_stocks(df):
    df = df.dropna(subset=["sharpe_6", "sharpe_3"]).copy()
    df["sharpe_6_rank"] = df["sharpe_6"].rank(ascending=False, method="first").astype(int)
    df["sharpe_3_rank"] = df["sharpe_3"].rank(ascending=False, method="first").astype(int)
    df["Avg_sharpe_6_3_Rank"] = df["sharpe_6_rank"] + df["sharpe_3_rank"]
    df.sort_values("Avg_sharpe_6_3_Rank", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def run_screener(
    mcap_filter=MCAP_FILTER_CR,
    roc_filter=ROC_ANNUAL_FILTER,
    turnover_filter=TURNOVER_FILTER_CR,
    as_of_date=None,
    long_months=6,
    short_months=3,
):
    if long_months <= 0 or short_months <= 0:
        raise ValueError("Sharpe month windows must be positive integers.")
    if long_months > 12 or short_months > 12:
        raise ValueError("Sharpe month windows cannot exceed 12.")

    t_start = time.time()
    long_days = months_to_trading_days(long_months)
    short_days = months_to_trading_days(short_months)
    days_to_load = max(DAYS_TO_LOAD, long_days + 1, short_days + 1)

    log.info("")
    log.info("=" * 60)
    log.info("NSE SHARPE RATIO SCREENER")
    log.info("=" * 60)
    log.info(
        f"  Mode            : {'HISTORICAL  (as of ' + as_of_date + ')' if as_of_date else 'LIVE  (latest data)'}"
    )
    log.info(f"  MCAP filter     : > Rs.{mcap_filter:,} Cr")
    log.info(f"  Annual ROC      : >= {roc_filter}%")
    log.info(f"  Median Turnover : >= Rs.{turnover_filter} Cr / day")
    log.info(f"  Sharpe windows  : {long_months}M and {short_months}M")
    log.info(f"  Sharpe formula  : mean(r) / std(r)  [no sqrt(252)]")
    log.info(f"  Price window    : last {days_to_load} trading days")
    log.info("")

    with get_connection() as conn:
        setup_schema(conn)
        snapshot_df, latest_date = load_mcap_snapshot(conn, as_of_date=as_of_date)
        mcap_df = apply_mcap_filter(snapshot_df, mcap_filter)
        if mcap_df.empty:
            return pd.DataFrame(), None
        prices_df = load_price_history(
            conn,
            mcap_df["symbol"].tolist(),
            days=days_to_load,
            as_of_date=as_of_date,
        )

    if prices_df.empty:
        log.warning("No price history found.")
        return pd.DataFrame(), None

    roc_df = compute_and_filter_roc(prices_df, roc_filter)
    if roc_df.empty:
        log.warning(f"No stocks passed Annual ROC >= {roc_filter}% filter.")
        return pd.DataFrame(), None

    turnover_df = compute_and_filter_turnover(
        prices_df, roc_df["symbol"].tolist(), turnover_filter
    )
    if turnover_df.empty:
        log.warning(
            f"No stocks passed Median Turnover >= Rs.{turnover_filter} Cr filter."
        )
        return pd.DataFrame(), None

    sharpe_df = compute_sharpe_vectorised(
        prices_df,
        turnover_df["symbol"].tolist(),
        long_months=long_months,
        short_months=short_months,
    )
    if sharpe_df.empty:
        log.warning("No stocks had sufficient history for Sharpe.")
        return pd.DataFrame(), None

    circuit_df = compute_circuit_hits(prices_df, sharpe_df["symbol"].tolist())

    log.info("Step 8 -- Merging and ranking ...")
    result = (
        mcap_df
        .merge(roc_df, on="symbol", how="inner")
        .merge(turnover_df, on="symbol", how="inner")
        .merge(sharpe_df, on="symbol", how="inner")
        .merge(circuit_df, on="symbol", how="left")
    )
    result["total_circuit_hits_3m"] = result["total_circuit_hits_3m"].fillna(0).astype(int)
    log.info(f"  Pool before ranking : {len(result):,} stocks")

    result = rank_stocks(result)
    result.attrs["long_months"] = int(long_months)
    result.attrs["short_months"] = int(short_months)

    output_cols = [
        "symbol", "company_name", "series",
        "close", "dma_20", "dma_50", "dma_100", "dma_200",
        "away_52wh",
        "Avg_sharpe_6_3_Rank",
        "sharpe_6", "sharpe_3",
        "ROC_6", "ROC_3",
        "week_52_high", "market_cap_cr", "ROC_annual",
        "median_turnover_cr",
        "sharpe_6_rank", "sharpe_3_rank",
        "total_circuit_hits_3m",
        "isin", "shares_outstanding",
    ]
    result = result[[c for c in output_cols if c in result.columns]]

    elapsed = time.time() - t_start
    log.info(f"  Final ranked list : {len(result):,} stocks")
    log.info(f"  Total time        : {elapsed:.1f}s")
    log.info("")
    return result, latest_date


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
FILTER_FILL = PatternFill("solid", start_color="1A5C38")
ALT_FILL2 = PatternFill("solid", start_color="D4EDDA")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=9)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    "symbol": 12,
    "company_name": 28,
    "series": 8,
    "close": 10,
    "dma_20": 10,
    "dma_50": 10,
    "dma_100": 10,
    "dma_200": 10,
    "away_52wh": 15,
    "Avg_sharpe_6_3_Rank": 16,
    "sharpe_6": 11,
    "sharpe_3": 11,
    "ROC_6": 10,
    "ROC_3": 10,
    "week_52_high": 13,
    "market_cap_cr": 14,
    "ROC_annual": 12,
    "median_turnover_cr": 16,
    "sharpe_6_rank": 13,
    "sharpe_3_rank": 13,
    "total_circuit_hits_3m": 18,
    "isin": 16,
    "shares_outstanding": 18,
}

FRIENDLY_HEADERS = {
    "symbol": "Symbol",
    "company_name": "Company Name",
    "series": "Series",
    "close": "Close (Rs.)",
    "dma_20": "20 DMA",
    "dma_50": "50 DMA",
    "dma_100": "100 DMA",
    "dma_200": "200 DMA",
    "away_52wh": "Away from 52WH %",
    "total_circuit_hits_3m": "Circuit Hits (3M)",
    "Avg_sharpe_6_3_Rank": "Avg Sharpe Rank",
    "sharpe_6": "Sharpe 6M",
    "sharpe_3": "Sharpe 3M",
    "ROC_6": "6M ROC %",
    "ROC_3": "3M ROC %",
    "week_52_high": "52W High (Rs.)",
    "market_cap_cr": "MCAP (Cr)",
    "ROC_annual": "Annual ROC %",
    "median_turnover_cr": "Med. Turnover (Cr)",
    "sharpe_6_rank": "Sharpe 6M Rank",
    "sharpe_3_rank": "Sharpe 3M Rank",
    "isin": "ISIN",
    "shares_outstanding": "Shares Outstanding",
}

NUM_COLS = {
    "close", "dma_20", "dma_50", "dma_100", "dma_200",
    "week_52_high", "market_cap_cr", "median_turnover_cr",
    "sharpe_6", "sharpe_3", "shares_outstanding",
}
PCT_COLS = {"ROC_annual", "ROC_3", "ROC_6", "away_52wh"}
INT_COLS = {"sharpe_6_rank", "sharpe_3_rank", "Avg_sharpe_6_3_Rank", "total_circuit_hits_3m"}


def _write_sheet(ws, df, title, header_fill, alt_fill, friendly_headers=None):
    cols = list(df.columns)
    today_str = datetime.today().strftime("%d %b %Y")
    headers = friendly_headers or FRIENDLY_HEADERS

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=f"{title}  |  Screened: {today_str}")
    title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_cell.fill = header_fill
    title_cell.alignment = CENTER

    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=headers.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(cols, start=1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            cell.fill = fill

            if col in PCT_COLS:
                cell.value = val / 100 if pd.notna(val) else None
                cell.number_format = "0.00%"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            elif col in NUM_COLS:
                cell.value = round(float(val), 2) if pd.notna(val) else None
                cell.number_format = "#,##0.00"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            elif col in INT_COLS:
                cell.value = int(val) if pd.notna(val) else None
                cell.number_format = "0"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.alignment = LEFT
                cell.font = DATA_FONT

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def apply_filtered_workbook(df):
    filtered = df.copy()

    if "ROC_3" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["ROC_3"].notna() &
            (filtered["ROC_3"] > ROC_3M_FILTER)
        ]
        log.info(
            f"  Filter a (ROC_3 > {ROC_3M_FILTER}%) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    if "week_52_high" in filtered.columns and "close" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["week_52_high"].notna() &
            (filtered["close"] >= filtered["week_52_high"] * 0.75)
        ]
        log.info(
            f"  Filter b (Close >= 52WH x 75%) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    if "total_circuit_hits_3m" in filtered.columns:
        before = len(filtered)
        filtered = filtered[
            filtered["total_circuit_hits_3m"] <= CIRCUIT_MAX_HITS
        ]
        log.info(
            f"  Filter c (Circuit hits <= {CIRCUIT_MAX_HITS}) : "
            f"{len(filtered):,} passed  (excluded {before - len(filtered):,})"
        )

    return filtered.reset_index(drop=True)


def export_to_excel(result, latest_date):
    base = Path(__file__).parent
    filename = f"{latest_date}.xlsx"
    out_path = base / filename
    long_months = int(result.attrs.get("long_months", 6))
    short_months = int(result.attrs.get("short_months", 3))

    friendly_headers = dict(FRIENDLY_HEADERS)
    friendly_headers["sharpe_6"] = f"Sharpe {long_months}M"
    friendly_headers["sharpe_3"] = f"Sharpe {short_months}M"
    friendly_headers["ROC_6"] = f"{long_months}M ROC %"
    friendly_headers["ROC_3"] = f"{short_months}M ROC %"
    friendly_headers["sharpe_6_rank"] = f"Sharpe {long_months}M Rank"
    friendly_headers["sharpe_3_rank"] = f"Sharpe {short_months}M Rank"
    friendly_headers["Avg_sharpe_6_3_Rank"] = f"Avg Sharpe {long_months}M/{short_months}M Rank"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Stocks"
    _write_sheet(
        ws1, result,
        title="NSE Sharpe Screener -- All Ranked Stocks",
        header_fill=HEADER_FILL,
        alt_fill=ALT_FILL,
        friendly_headers=friendly_headers,
    )
    log.info(f"  Sheet 1 written : All Stocks  ({len(result):,} rows)")

    filtered = apply_filtered_workbook(result)
    ws2 = wb.create_sheet(title="Filtered")
    if filtered.empty:
        ws2.cell(row=1, column=1, value="No stocks matched the filter criteria.")
        log.warning("  No stocks passed filtered sheet criteria.")
    else:
        _write_sheet(
            ws2, filtered,
            title=(
                f"Filtered  |  3M ROC > {ROC_3M_FILTER}%  |  "
                f"Close >= 52WH x 75%  |  Circuit Hits <= {CIRCUIT_MAX_HITS}"
            ),
            header_fill=FILTER_FILL,
            alt_fill=ALT_FILL2,
            friendly_headers=friendly_headers,
        )
        log.info(f"  Sheet 2 written : Filtered  ({len(filtered):,} rows)")

    wb.save(str(out_path))
    log.info("-" * 50)
    log.info(f"SUCCESS: Sharpe Screener output saved to Excel.")
    log.info(f"File Path: {out_path.absolute()}")
    log.info("-" * 50)


def print_results(df, roc_filter=ROC_ANNUAL_FILTER,
                  turnover_filter=TURNOVER_FILTER_CR):
    if df.empty:
        print("No results to display.")
        return

    long_months = int(df.attrs.get("long_months", 6))
    short_months = int(df.attrs.get("short_months", 3))
    display = df.copy()
    display.index = range(1, len(display) + 1)

    for col in ["ROC_annual", "ROC_6", "ROC_3", "away_52wh"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:+.2f}%" if x is not None and str(x) != "nan" else ""
            )
    for col in ["sharpe_6", "sharpe_3"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:+.4f}" if x is not None and str(x) != "nan" else ""
            )
    for col in ["close", "dma_20", "dma_50", "dma_100", "dma_200"]:
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"{x:,.2f}" if x is not None and str(x) != "nan" else ""
            )
    if "market_cap_cr" in display.columns:
        display["market_cap_cr"] = display["market_cap_cr"].apply(
            lambda x: f"Rs.{x:,.0f} Cr" if x == x and x is not None else "N/A"
        )
    if "median_turnover_cr" in display.columns:
        display["median_turnover_cr"] = display["median_turnover_cr"].apply(
            lambda x: f"Rs.{x:.2f} Cr" if x == x and x is not None else "N/A"
        )

    cols = [
        "symbol", "company_name", "market_cap_cr",
        "ROC_annual", "median_turnover_cr",
        "sharpe_6", "sharpe_6_rank",
        "sharpe_3", "sharpe_3_rank",
        "Avg_sharpe_6_3_Rank",
    ]
    cols = [c for c in cols if c in display.columns]

    print(f"\n{'=' * 100}")
    print(
        f"  SHARPE RATIO SCREENER RESULTS  |  "
        f"Screened: {datetime.today().strftime('%d %b %Y')}"
    )
    print(f"{'=' * 100}")
    print(display[cols].to_string())
    print(f"{'=' * 100}")
    print(
        f"  Filters : MCAP > Rs.{MCAP_FILTER_CR:,} Cr  |  "
        f"Annual ROC >= {roc_filter}%  |  "
        f"Median Turnover >= Rs.{turnover_filter} Cr"
    )
    print(f"  Windows : Sharpe {long_months}M and {short_months}M")
    print(f"  Total Stocks: {len(df)}")
    print("  Formula : Sharpe = mean(daily_returns) / std(daily_returns)")
    print("  Note    : Median turnover used (not mean) to ignore block deal spikes")
    print("            sqrt(252) omitted -- ranking is identical\n")


def main():
    args = sys.argv[1:]
    mcap_filter = MCAP_FILTER_CR
    roc_filter = ROC_ANNUAL_FILTER
    turnover_filter = TURNOVER_FILTER_CR
    top_n = TOP_N
    as_of_date = None
    long_months = 6
    short_months = 3

    for i, arg in enumerate(args):
        if arg == "--mcap" and i + 1 < len(args):
            try:
                mcap_filter = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--top" and i + 1 < len(args):
            try:
                top_n = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--rf" and i + 1 < len(args):
            try:
                roc_filter = float(args[i + 1])
            except ValueError:
                pass
        if arg == "--turnover" and i + 1 < len(args):
            try:
                turnover_filter = float(args[i + 1])
            except ValueError:
                pass
        if arg == "--date" and i + 1 < len(args):
            raw = args[i + 1].strip()
            try:
                datetime.strptime(raw, "%Y-%m-%d")
                as_of_date = raw
            except ValueError:
                print(f"Invalid --date format: '{raw}'. Use YYYY-MM-DD.")
                return
        if arg == "--long-months" and i + 1 < len(args):
            try:
                long_months = int(args[i + 1])
            except ValueError:
                pass
        if arg == "--short-months" and i + 1 < len(args):
            try:
                short_months = int(args[i + 1])
            except ValueError:
                pass

    if long_months <= 0 or short_months <= 0:
        print("Invalid Sharpe month values. Use positive integers.")
        return
    if long_months > 12 or short_months > 12:
        print("Invalid Sharpe month values. Maximum allowed is 12.")
        return

    result, latest_date = run_screener(
        mcap_filter=mcap_filter,
        roc_filter=roc_filter,
        turnover_filter=turnover_filter,
        as_of_date=as_of_date,
        long_months=long_months,
        short_months=short_months,
    )
    if result.empty or latest_date is None:
        return

    print_results(
        result,
        roc_filter=roc_filter,
        turnover_filter=turnover_filter,
    )
    export_to_excel(result, latest_date)


if __name__ == "__main__":
    main()
